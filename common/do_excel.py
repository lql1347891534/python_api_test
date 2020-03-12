import openpyxl
from common.api_request import ApiRequest
import json

class Case:

    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.expected=None
        self.actual=None
        self.result=None

class ReadingExcel:

    def __init__(self,file_name):
        try:
            self.workbook=openpyxl.load_workbook(filename=file_name)
        except FileNotFoundError as e:
            print("{0} not found ,please check file path".format(file_name))
            raise e

    def get_data(self,sheet_name):
        sheet=self.workbook[sheet_name]
        cases=[]#创建一个列表，存储Excel中提取的测试数据
        max_row=sheet.max_row #获取最大行数
        for r in range(2,max_row+1):
            case=Case()#实例化一个case对象，用来存储Excel中获取的测试用例
            case.case_id=sheet.cell(row=r,column=1).value
            case.title = sheet.cell(row=r, column=2).value
            case.url = sheet.cell(row=r, column=3).value
            case.data = sheet.cell(row=r, column=4).value
            case.method = sheet.cell(row=r, column=5).value
            case.expected = sheet.cell(row=r, column=6).value
            cases.append(case)#将获取到的测试数据添加到列表中

        return cases

    def get_sheet_name(self):
        return self.workbook.sheetnames

if __name__ == '__main__':
    do_excel=ReadingExcel(r"D:\Python_api_test\\testcases\\testcase.xlsx")
    cases=ReadingExcel(r"D:\Python_api_test\\testcases\\testcase.xlsx").get_data("login")

    for case in cases:
        data=json.loads(case.data)#反序列化将从Excel中获取的str类型的数据转换成Python可以识别的dict类型
        resp=ApiRequest(method=case.method,url=case.url,data=data)
        print(resp.get_status_code())
        print(resp.get_json())
        print(resp.get_text())











